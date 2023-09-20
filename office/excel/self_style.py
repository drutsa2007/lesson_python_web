from openpyxl import load_workbook
from openpyxl.styles import NamedStyle, Font, Border, Side

wb = load_workbook(filename='9.xlsx')
ws = wb.active

# Создание стиля
name_style = NamedStyle(name="highlight")
name_style.font = Font(bold=True, size=20)
bd = Side(style='thick', color="000000")
name_style.border = Border(left=bd, top=bd, right=bd, bottom=bd)

wb.add_named_style(name_style)

ws['A1'].style = name_style
# или
ws['D5'].style = 'highlight'

wb.save('9.xlsx')
