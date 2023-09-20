from openpyxl import Workbook

new_wb = Workbook()

# текущий лист
sheet = new_wb.active

sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
sheet.page_setup.paperSize = sheet.PAPERSIZE_TABLOID
sheet.page_setup.fitToHeight = 0
sheet.page_setup.fitToWidth = 1

# переименовываем активный лист
sheet.title = "New_sheet2"
# добавляем новый лист
new_sheet = new_wb.create_sheet('New Sheet3')

# добавляем данные в активный лист
sheet.append([1, 2, 3, 4])
sheet['E1'] = "=SUM(A1:D1)"
sheet['E3'] = "Текст"

# вставляем новую строку и столбец
sheet.insert_rows(2)
sheet.insert_cols(2)

# удаляем строку и столбец
sheet.delete_rows(2)
sheet.delete_cols(2)

# объединяем ячейки
sheet.merge_cells("A1:A2")
# разъединяем ячейки
sheet.unmerge_cells("A1:A2")

# работа с ячейкой
cell = sheet['B7']  # либо дипазон
cell.value = "Данные {}".format(sheet.max_row)  # Пишем формулу в ячейку

new_wb.save('new_book.xlsx')