from openpyxl import load_workbook


wb = load_workbook(filename='9.xlsx')

# Получить имена листов
print(wb.sheetnames)

ws = wb['Лист1']
print(ws.title)
c = ws['A3']
print(c.row, c.column, c.coordinate)


# получение данных из ячеек
print(ws['D18'].value)
for i in ws['A1:A4']:
    print(i[0].value)

for i in ws['A1':'A4']:
    print(i[0].value)
# с помощью команды dir можно вытащить другую информацию из ячейки
# print(dir(ws['A1':'A4'][0][0]))

for row in ws.iter_rows(min_row=1, max_col=3, max_row=2, values_only=True):
    print(row)
