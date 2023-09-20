import datetime

from openpyxl import load_workbook
from openpyxl.styles import Font, Color, Alignment, colors, Side, Border, PatternFill

wb = load_workbook(filename='9.xlsx')
ws = wb.active
cell = ws['A1']

cell.font = Font(color="FF0000", bold=True)
# Параметры Font: 'b', 'bold', 'charset', 'color', 'condense', 'extend', 'family', 'from_tree', 'i', 'idx_base',
# 'italic', 'name', 'namespace', 'outline', 'scheme', 'shadow', 'size', 'strike', 'strikethrough', 'sz', 'tagname',
# 'to_tree', 'u', 'underline', 'vertAlign'

my_align = Alignment(vertical="center", horizontal="center")
cell.alignment = my_align

# Высота строки
ws.row_dimensions[2].height = 30
# Ширина столбца
ws.column_dimensions['B'].width = 40

# Границы ячейки
thins = Side(border_style="medium", color="0000ff")
double = Side(border_style="dashDot", color="ff0000")
ws['B2'].border = Border(top=double, bottom=double, left=thins, right=thins)

# Заливка цветом
ws.merge_cells('C3:D7')
merge_cell = ws['C3']
merge_cell.fill = PatternFill('solid', fgColor="d4ef7a")

# Форматы ячейки
ws['A1'] = datetime.datetime(2010, 7, 21)
ws['A1'].number_format = 'yyyy-mm-dd h:mm:ss'
ws["A2"] = 0.123456
ws["A2"].number_format = "0.00"

wb.save('9.xlsx')
